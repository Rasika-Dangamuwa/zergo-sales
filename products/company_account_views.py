"""
Company Account Management Views
Handles company accounts, transactions, opening balances, and ledger
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Q, F
from django.utils import timezone
from decimal import Decimal
import json

from .models import Company, CompanyAccount, CompanyTransaction, Purchase, PurchaseReturn, CompanyPayment, PaymentAllocation


@login_required
def company_account_list(request):
    """List all company accounts with balances"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied. Only admin and office staff can view company accounts.')
        return redirect('dashboard:home')
    
    accounts = CompanyAccount.objects.select_related('company').order_by('company__company_name')
    
    # Calculate summary stats
    total_payable = accounts.filter(current_balance__gt=0).aggregate(total=Sum('current_balance'))['total'] or 0
    total_receivable = accounts.filter(current_balance__lt=0).aggregate(total=Sum('current_balance'))['total'] or 0
    
    context = {
        'accounts': accounts,
        'total_payable': total_payable,
        'total_receivable': abs(total_receivable),
    }
    return render(request, 'products/company_account_list.html', context)


@login_required
def company_account_detail(request, pk):
    """World-class company account detail with advanced analytics"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied. Only admin and office staff can view company accounts.')
        return redirect('dashboard:home')
    
    account = get_object_or_404(CompanyAccount, pk=pk)
    
    # Get transactions
    transactions = account.transactions.select_related(
        'purchase', 'purchase_return', 'created_by'
    ).order_by('-transaction_date', '-created_at')
    
    # Filter by date range if provided
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    transaction_type_filter = request.GET.get('transaction_type')
    search_query = request.GET.get('search')
    
    # Apply filters
    filtered_transactions = transactions
    if from_date:
        filtered_transactions = filtered_transactions.filter(transaction_date__gte=from_date)
    if to_date:
        filtered_transactions = filtered_transactions.filter(transaction_date__lte=to_date)
    if transaction_type_filter:
        filtered_transactions = filtered_transactions.filter(transaction_type=transaction_type_filter)
    if search_query:
        filtered_transactions = filtered_transactions.filter(
            Q(reference_number__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(notes__icontains=search_query)
        )
    
    # Calculate summary statistics for period
    from decimal import Decimal
    from datetime import timedelta
    
    total_purchases = filtered_transactions.filter(transaction_type='purchase').aggregate(
        total=Sum('amount'))['total'] or Decimal('0')
    total_returns = filtered_transactions.filter(transaction_type='return').aggregate(
        total=Sum('amount'))['total'] or Decimal('0')
    total_payments = filtered_transactions.filter(transaction_type='payment').aggregate(
        total=Sum('amount'))['total'] or Decimal('0')
    
    # Outstanding GRNs Analysis - Get all unpaid/partially paid purchases
    today = timezone.localdate()
    all_purchases = Purchase.objects.filter(
        company=account.company,
        status='received'
    ).select_related('created_by').order_by('-grn_date')
    
    # Filter to only those with actual outstanding amounts (using property calculation)
    outstanding_purchases_list = []
    for purchase in all_purchases:
        if purchase.amount_outstanding > Decimal('0.01'):  # More than 1 paisa outstanding
            purchase.days_outstanding = (today - purchase.grn_date.date()).days
            outstanding_purchases_list.append(purchase)
            if len(outstanding_purchases_list) >= 10:  # Limit to top 10
                break
    
    outstanding_purchases = outstanding_purchases_list
    total_outstanding_amount = sum(p.amount_outstanding for p in outstanding_purchases)
    
    # Aging Analysis (30/60/90 days) - Calculate based on actual outstanding amounts
    aging_30 = Decimal('0')
    aging_60 = Decimal('0')
    aging_90 = Decimal('0')
    aging_90_plus = Decimal('0')
    
    for purchase in Purchase.objects.filter(
        company=account.company,
        status='received'
    ).select_related('created_by'):
        outstanding = purchase.amount_outstanding
        
        # Only include purchases with actual outstanding amounts
        if outstanding > Decimal('0.01'):
            days_old = (today - purchase.grn_date.date()).days
            
            if days_old <= 30:
                aging_30 += outstanding
            elif days_old <= 60:
                aging_60 += outstanding
            elif days_old <= 90:
                aging_90 += outstanding
            else:
                aging_90_plus += outstanding
    
    # Payment Performance Metrics
    paid_purchases = Purchase.objects.filter(
        company=account.company,
        status='received',
        payment_status='paid'
    )
    
    # Calculate average payment cycle (days from GRN to full payment)
    payment_cycles = []
    for purchase in paid_purchases:
        # Get last payment allocation for this purchase
        last_payment = purchase.payment_allocations.order_by('-payment__payment_date').first()
        if last_payment:
            cycle_days = (last_payment.payment.payment_date.date() - purchase.grn_date.date()).days
            if cycle_days >= 0:  # Exclude negative values
                payment_cycles.append(cycle_days)
    
    avg_payment_cycle = sum(payment_cycles) / len(payment_cycles) if payment_cycles else 0
    
    # Transaction trends (last 6 months)
    six_months_ago = today - timedelta(days=180)
    monthly_trends = []
    
    from calendar import monthrange
    for i in range(6):
        month_start = (today.replace(day=1) - timedelta(days=30*i)).replace(day=1)
        month_end = month_start.replace(day=monthrange(month_start.year, month_start.month)[1])
        
        month_txns = account.transactions.filter(
            transaction_date__date__gte=month_start,
            transaction_date__date__lte=month_end
        )
        
        month_purchases = month_txns.filter(transaction_type='purchase').aggregate(
            total=Sum('amount'))['total'] or Decimal('0')
        month_payments = month_txns.filter(transaction_type='payment').aggregate(
            total=Sum('amount'))['total'] or Decimal('0')
        
        monthly_trends.insert(0, {
            'month': month_start.strftime('%b %Y'),
            'purchases': float(month_purchases),
            'payments': float(abs(month_payments)),
        })
    
    # Calculate running balance and get settlement details
    running_transactions = []
    balance = account.opening_balance
    
    for txn in filtered_transactions:
        # Get settlement details if this is a purchase return
        settlement_details = []
        outstanding = None
        payment_allocations = []
        days_outstanding = None
        
        if txn.purchase_return:
            # Get all settlement records for this return
            from products.models import PurchaseReturnSettlement
            settlements = PurchaseReturnSettlement.objects.filter(
                purchase_return=txn.purchase_return
            ).select_related('replacement_grn')
            settlement_details = list(settlements)
        
        # Get outstanding amount and payment history if this is a purchase
        if txn.purchase:
            outstanding = txn.purchase.amount_outstanding
            payment_allocations = txn.purchase.payment_allocations.select_related('payment').all()
            if outstanding > 0:
                days_outstanding = (today - txn.purchase.grn_date.date()).days
        
        # Calculate balance
        # All transaction amounts already have correct signs:
        # - Purchases/debits: positive (increase what we owe)
        # - Returns/payments/credits: negative (decrease what we owe)
        # So we just add all amounts - negative amounts auto-reduce balance
        balance += txn.amount
        
        running_transactions.append({
            'transaction': txn,
            'balance_after': balance,
            'settlement_details': settlement_details,
            'outstanding': outstanding,
            'payment_allocations': payment_allocations,
            'days_outstanding': days_outstanding,
        })
    
    context = {
        'account': account,
        'transactions': running_transactions,
        'from_date': from_date,
        'to_date': to_date,
        'transaction_type_filter': transaction_type_filter,
        'search_query': search_query,
        
        # Summary stats
        'total_purchases': total_purchases,
        'total_returns': abs(total_returns),
        'total_payments': abs(total_payments),
        'transaction_count': filtered_transactions.count(),
        
        # Outstanding GRNs
        'outstanding_purchases': outstanding_purchases,
        'total_outstanding_amount': total_outstanding_amount,
        'outstanding_count': len(outstanding_purchases),
        
        # Aging analysis
        'aging_30': aging_30,
        'aging_60': aging_60,
        'aging_90': aging_90,
        'aging_90_plus': aging_90_plus,
        
        # Performance metrics
        'avg_payment_cycle': round(avg_payment_cycle, 1),
        'paid_purchases_count': paid_purchases.count(),
        
        # Trends
        'monthly_trends': monthly_trends,
    }
    return render(request, 'products/company_account_detail.html', context)


@login_required
def create_opening_balance(request):
    """Create or update opening balance for a company"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied. Only admin and office staff can create opening balances.')
        return redirect('dashboard:home')
    
    companies = Company.objects.filter(is_active=True).order_by('company_name')
    
    if request.method == 'POST':
        try:
            company_id = request.POST.get('company')
            opening_balance = Decimal(request.POST.get('opening_balance', '0'))
            opening_date = request.POST.get('opening_date')
            opening_notes = request.POST.get('opening_notes', '')
            
            company = Company.objects.get(pk=company_id)
            
            # Get or create account
            account, created = CompanyAccount.objects.get_or_create(
                company=company,
                defaults={
                    'opening_balance': opening_balance,
                    'current_balance': opening_balance,
                    'opening_date': opening_date,
                    'opening_notes': opening_notes,
                    'created_by': request.user,
                }
            )
            
            if not created:
                # Update existing account
                account.opening_balance = opening_balance
                account.opening_date = opening_date
                account.opening_notes = opening_notes
                account.save()
                
                # Recalculate current balance
                account.update_balance()
                
                messages.success(request, f'Opening balance updated for {company.company_name}')
            else:
                messages.success(request, f'Opening balance created for {company.company_name}')
            
            return redirect('products:company_account_detail', pk=account.pk)
            
        except Exception as e:
            messages.error(request, f'Error creating opening balance: {str(e)}')
    
    context = {
        'companies': companies,
    }
    return render(request, 'products/create_opening_balance.html', context)


@login_required
def record_company_payment(request):
    """Record a payment to company (cash, cheque, bank transfer)"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    companies = Company.objects.filter(is_active=True).order_by('company_name')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                company_id = request.POST.get('company')
                amount = Decimal(request.POST.get('amount', '0'))
                payment_date = request.POST.get('payment_date')
                settlement_method = request.POST.get('settlement_method')
                payment_reference = request.POST.get('payment_reference', '')
                notes = request.POST.get('notes', '')
                
                company = Company.objects.get(pk=company_id)
                account = CompanyAccount.objects.get(company=company)
                
                # Create transaction
                CompanyTransaction.objects.create(
                    company_account=account,
                    transaction_type='payment',
                    transaction_date=timezone.now(),
                    reference_number=payment_reference or f'PAY-{timezone.now().strftime("%Y%m%d-%H%M%S")}',
                    amount=-amount,
                    settlement_method=settlement_method,
                    payment_reference=payment_reference,
                    description=f'Payment via {settlement_method}',
                    notes=notes,
                    created_by=request.user,
                )
                
                messages.success(request, f'Payment of Rs. {amount:,.2f} recorded for {company.company_name}')
                return redirect('products:company_account_detail', pk=account.pk)
                
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')
    
    context = {
        'companies': companies,
    }
    return render(request, 'products/record_company_payment.html', context)


@login_required
def settle_grn_with_return(request):
    """Settle a GRN/purchase with a return (offset)"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                purchase_id = request.POST.get('purchase')
                return_id = request.POST.get('return')
                settlement_amount = Decimal(request.POST.get('settlement_amount', '0'))
                
                purchase = Purchase.objects.get(pk=purchase_id)
                purchase_return = PurchaseReturn.objects.get(pk=return_id)
                
                if purchase.company != purchase_return.company:
                    messages.error(request, 'Purchase and return must be from the same company')
                    return redirect('products:company_account_list')
                
                account = CompanyAccount.objects.get(company=purchase.company)
                
                # Create settlement transaction
                settlement_txn = CompanyTransaction.objects.create(
                    company_account=account,
                    transaction_type='adjustment',
                    transaction_date=timezone.now(),
                    reference_number=f'SETTLE-{purchase.grn_number}-{purchase_return.pr_number}',
                    amount=0,
                    settlement_method='grn_offset',
                    purchase=purchase,
                    purchase_return=purchase_return,
                    description=f'Settlement: GRN {purchase.grn_number} vs Return {purchase_return.pr_number}',
                    notes=f'Offset settlement - GRN amount: Rs.{purchase.total_amount}, Return amount: Rs.{purchase_return.total_amount}',
                    created_by=request.user,
                )
                
                messages.success(request, f'Settlement of Rs. {settlement_amount:,.2f} recorded')
                return redirect('products:company_account_detail', pk=account.pk)
                
        except Exception as e:
            messages.error(request, f'Error creating settlement: {str(e)}')
    
    return redirect('products:company_account_list')


@login_required
def record_company_payment(request):
    """Record payment made to company with GRN allocation"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                company_id = request.POST.get('company')
                payment_method = request.POST.get('payment_method')
                total_amount = Decimal(request.POST.get('total_amount', '0'))
                payment_date = request.POST.get('payment_date') or timezone.now()
                
                # Method-specific details
                cheque_number = request.POST.get('cheque_number', '')
                cheque_date = request.POST.get('cheque_date') or None
                bank_name = request.POST.get('bank_name', '')
                transfer_reference = request.POST.get('transfer_reference', '')
                transfer_date = request.POST.get('transfer_date') or None
                notes = request.POST.get('notes', '')
                
                company = Company.objects.get(pk=company_id)
                
                # Create payment record
                payment = CompanyPayment.objects.create(
                    company=company,
                    payment_method=payment_method,
                    total_amount=total_amount,
                    payment_date=payment_date,
                    cheque_number=cheque_number if payment_method == 'cheque' else None,
                    cheque_date=cheque_date if payment_method == 'cheque' else None,
                    bank_name=bank_name if payment_method in ['cheque', 'bank_transfer'] else None,
                    transfer_reference=transfer_reference if payment_method == 'bank_transfer' else None,
                    transfer_date=transfer_date if payment_method == 'bank_transfer' else None,
                    reference_notes=notes,
                    created_by=request.user
                )
                
                # Create allocations to GRNs
                grn_ids = request.POST.getlist('grn_ids[]')
                allocation_amounts = request.POST.getlist('allocation_amounts[]')
                
                total_allocated = Decimal('0')
                allocations_created = []
                
                for grn_id, amount_str in zip(grn_ids, allocation_amounts):
                    if not amount_str or Decimal(amount_str) == 0:
                        continue
                    
                    allocated_amount = Decimal(amount_str)
                    
                    # VALIDATION: Check cumulative total doesn't exceed payment amount
                    new_total = total_allocated + allocated_amount
                    if new_total > total_amount:
                        raise ValueError(
                            f'Cannot allocate Rs. {allocated_amount:,.2f} to next GRN. '
                            f'Total allocations would be Rs. {new_total:,.2f} but payment is only Rs. {total_amount:,.2f}. '
                            f'Already allocated: Rs. {total_allocated:,.2f}'
                        )
                    
                    purchase = Purchase.objects.get(pk=grn_id)
                    
                    # Validate allocation doesn't exceed GRN outstanding balance
                    if allocated_amount > purchase.amount_outstanding:
                        raise ValueError(
                            f'Cannot allocate Rs. {allocated_amount:,.2f} to {purchase.grn_number}. '
                            f'Outstanding balance is only Rs. {purchase.amount_outstanding:,.2f} '
                            f'(Total: Rs. {purchase.total_amount:,.2f}, '
                            f'Already Paid: Rs. {purchase.total_paid:,.2f}, '
                            f'Settled via Returns: Rs. {purchase.total_settled_via_returns:,.2f})'
                        )
                    
                    PaymentAllocation.objects.create(
                        payment=payment,
                        purchase=purchase,
                        allocated_amount=allocated_amount,
                        created_by=request.user
                    )
                    purchase.sync_payment_status()  # Sync stored payment_status / amount_paid
                    total_allocated += allocated_amount
                    allocations_created.append(f'{purchase.grn_number}: Rs. {allocated_amount:,.2f}')
                
                # Validate total allocated matches payment amount
                if total_allocated != total_amount:
                    raise ValueError(
                        f'Total allocated (Rs. {total_allocated:,.2f}) does not match '
                        f'payment amount (Rs. {total_amount:,.2f}). '
                        f'Please adjust allocations.'
                    )
                
                messages.success(
                    request,
                    f'Payment {payment.payment_number} of Rs. {total_amount:,.2f} recorded successfully. '
                    f'Allocated to {len(allocations_created)} GRN(s).'
                )
                return redirect('products:payment_detail', pk=payment.pk)
                
        except Exception as e:
            messages.error(request, f'Error recording payment: {str(e)}')
            return redirect('products:record_company_payment')
    
    # GET request - show form
    companies = Company.objects.filter(is_active=True).order_by('company_name')
    
    # Get URL parameters for pre-selection
    selected_company_id = request.GET.get('company')
    selected_grn_id = request.GET.get('grn')
    
    context = {
        'companies': companies,
        'selected_company_id': selected_company_id,
        'selected_grn_id': selected_grn_id,
    }
    return render(request, 'products/record_company_payment.html', context)


@login_required
def get_company_outstanding_grns(request, company_id):
    """AJAX endpoint to get outstanding GRNs for a company"""
    import json
    from django.http import JsonResponse
    
    company = get_object_or_404(Company, pk=company_id)
    
    # Get GRNs with outstanding balance
    grns = Purchase.objects.filter(
        company=company,
        status='received',
        stock_updated=True
    ).select_related('company').order_by('-grn_date')
    
    grn_list = []
    for grn in grns:
        outstanding = grn.amount_outstanding
        if outstanding > 0:  # Only include GRNs with outstanding balance
            grn_list.append({
                'id': grn.pk,
                'grn_number': grn.grn_number,
                'grn_date': grn.grn_date.strftime('%Y-%m-%d'),
                'total_amount': float(grn.total_amount),
                'paid_amount': float(grn.total_paid),
                'returns_amount': float(grn.total_settled_via_returns),
                'outstanding': float(outstanding),
                'payment_status': grn.calculated_payment_status,
            })
    
    return JsonResponse({'grns': grn_list})


@login_required
def payment_detail(request, pk):
    """View payment details and allocations"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    payment = get_object_or_404(
        CompanyPayment.objects.select_related('company', 'created_by'),
        pk=pk
    )
    
    allocations = payment.allocations.select_related('purchase').order_by('-created_at')
    
    context = {
        'payment': payment,
        'allocations': allocations,
    }
    return render(request, 'products/payment_detail.html', context)


@login_required
def export_company_ledger(request, pk):
    """Export company account ledger to Excel"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    account = get_object_or_404(CompanyAccount, pk=pk)
    
    # Get transactions with filters
    transactions = account.transactions.select_related(
        'purchase', 'purchase_return', 'created_by'
    ).order_by('transaction_date', 'created_at')
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    
    if from_date:
        transactions = transactions.filter(transaction_date__gte=from_date)
    if to_date:
        transactions = transactions.filter(transaction_date__lte=to_date)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{account.company.company_name[:25]} Ledger"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = f"Company Account Ledger - {account.company.company_name}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Account Info
    ws['A2'] = f"Opening Balance: Rs. {account.opening_balance:,.2f}"
    ws['A3'] = f"Current Balance: Rs. {account.current_balance:,.2f}"
    ws['A4'] = f"Period: {from_date or 'All'} to {to_date or 'All'}"
    
    # Headers
    headers = ['Date', 'Type', 'Reference', 'Method', 'Debit', 'Credit', 'Balance', 'Outstanding', 'Notes']
    row_num = 6
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Opening Balance Row
    row_num += 1
    ws.cell(row=row_num, column=1).value = account.opening_date.strftime('%Y-%m-%d') if account.opening_date else ''
    ws.cell(row=row_num, column=2).value = 'Opening Balance'
    ws.cell(row=row_num, column=7).value = float(account.opening_balance)
    ws.cell(row=row_num, column=9).value = account.opening_notes or ''
    
    # Transaction Rows
    balance = account.opening_balance
    
    for txn in transactions:
        # Main transaction row
        row_num += 1
        
        ws.cell(row=row_num, column=1).value = txn.transaction_date.strftime('%Y-%m-%d')
        ws.cell(row=row_num, column=2).value = txn.get_transaction_type_display()
        
        # Reference
        if txn.purchase:
            ws.cell(row=row_num, column=3).value = txn.purchase.grn_number
        elif txn.purchase_return:
            ws.cell(row=row_num, column=3).value = txn.purchase_return.pr_number
        else:
            ws.cell(row=row_num, column=3).value = txn.payment_reference or ''
        
        ws.cell(row=row_num, column=4).value = txn.get_settlement_method_display()
        
        # Debit/Credit
        if txn.transaction_type in ['purchase', 'debit']:
            ws.cell(row=row_num, column=5).value = float(txn.amount)
            balance += txn.amount
        elif txn.transaction_type in ['return', 'payment', 'credit', 'settlement']:
            ws.cell(row=row_num, column=6).value = float(txn.amount)
            balance -= txn.amount
        
        ws.cell(row=row_num, column=7).value = float(balance)
        
        # Outstanding
        if txn.purchase:
            ws.cell(row=row_num, column=8).value = float(txn.purchase.amount_outstanding)
        
        ws.cell(row=row_num, column=9).value = txn.notes or ''
        
        # Settlement details sub-rows
        if txn.purchase_return:
            from products.models import PurchaseReturnSettlement
            settlements = PurchaseReturnSettlement.objects.filter(
                purchase_return=txn.purchase_return
            ).select_related('replacement_grn')
            
            for settlement in settlements:
                row_num += 1
                ws.cell(row=row_num, column=2).value = f"  ↪ {settlement.get_settlement_method_display()}"
                
                if settlement.replacement_grn:
                    ws.cell(row=row_num, column=3).value = settlement.replacement_grn.grn_number
                elif settlement.credit_note_number:
                    ws.cell(row=row_num, column=3).value = settlement.credit_note_number
                else:
                    ws.cell(row=row_num, column=3).value = settlement.refund_reference or 'Cash'
                
                ws.cell(row=row_num, column=6).value = float(settlement.settlement_amount)
                ws.cell(row=row_num, column=9).value = 'Settlement detail'
    
    # Format columns
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 30
    
    # Number formatting
    for row in range(7, row_num + 1):
        for col in [5, 6, 7, 8]:  # Debit, Credit, Balance, Outstanding
            cell = ws.cell(row=row, column=col)
            if cell.value:
                cell.number_format = '#,##0.00'
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Ledger_{account.company.company_name.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def payment_list(request):
    """List all company payments"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    payments = CompanyPayment.objects.select_related('company', 'created_by').order_by('-payment_date', '-payment_number')
    
    # Filter by company if provided
    company_id = request.GET.get('company')
    if company_id:
        payments = payments.filter(company_id=company_id)
    
    # Filter by payment method
    payment_method = request.GET.get('payment_method')
    if payment_method:
        payments = payments.filter(payment_method=payment_method)
    
    # Filter by date range
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date:
        payments = payments.filter(payment_date__gte=from_date)
    if to_date:
        payments = payments.filter(payment_date__lte=to_date)
    
    companies = Company.objects.filter(is_active=True).order_by('company_name')
    
    context = {
        'payments': payments,
        'companies': companies,
        'selected_company': company_id,
        'selected_method': payment_method,
        'from_date': from_date,
        'to_date': to_date,
    }
    return render(request, 'products/payment_list.html', context)
