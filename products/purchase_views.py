"""
Purchase/GRN Management Views
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from django.db.models import Sum, Count, Q, Case, When, Value, IntegerField
from django.utils import timezone
from django.http import HttpResponse
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime

from .models import (
    Purchase, PurchaseItem, PurchaseReturn, PurchaseReturnItem, 
    Product, Company, StockMovement, PurchaseOrder, PurchaseReturnSettlement,
    CompanyAccount, CompanyTransaction
)


@login_required
def purchase_list(request):
    """List all GRNs with advanced filtering"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied. Only admin and office staff can view purchases.')
        return redirect('dashboard:home')
    
    # Check for export request
    if request.GET.get('export') == 'excel':
        return export_purchases_excel(request)
    
    purchases = Purchase.objects.select_related('company', 'created_by').all()
    
    # Search functionality
    search = request.GET.get('search', '').strip()
    if search:
        purchases = purchases.filter(
            Q(grn_number__icontains=search) |
            Q(supplier_invoice_number__icontains=search) |
            Q(company__company_name__icontains=search)
        )
    
    # Date range filter
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date:
        purchases = purchases.filter(grn_date__date__gte=from_date)
    if to_date:
        purchases = purchases.filter(grn_date__date__lte=to_date)
    
    # Filter by status
    status = request.GET.get('status')
    if status:
        purchases = purchases.filter(status=status)
    
    # Filter by company
    company_id = request.GET.get('company')
    if company_id:
        purchases = purchases.filter(company_id=company_id)
    
    # Filter by payment status (using calculated_payment_status property)
    payment_status_filter = request.GET.get('payment_status')
    if payment_status_filter:
        # Get all purchases first, then filter by calculated property
        filtered_purchases = []
        for purchase in purchases:
            if purchase.calculated_payment_status == payment_status_filter:
                filtered_purchases.append(purchase.pk)
        purchases = purchases.filter(pk__in=filtered_purchases)
    
    # Filter by stock updated status
    stock_updated = request.GET.get('stock_updated')
    if stock_updated == 'yes':
        purchases = purchases.filter(stock_updated=True)
    elif stock_updated == 'no':
        purchases = purchases.filter(stock_updated=False)
    
    # Sorting
    sort = request.GET.get('sort', '-grn_date')
    purchases = purchases.order_by(sort)
    
    # Calculate comprehensive stats (before pagination)
    all_purchases = Purchase.objects.all()
    
    # Calculate accurate totals using properties (includes payments AND return settlements)
    paid_total = Decimal('0')
    outstanding_total = Decimal('0')
    unpaid_count = 0
    
    for purchase in all_purchases:
        paid_total += purchase.total_paid + purchase.total_settled_via_returns
        outstanding_total += purchase.amount_outstanding
        
        # Use calculated_payment_status for accurate unpaid count
        if purchase.calculated_payment_status == 'unpaid':
            unpaid_count += 1
    
    stats = {
        'total_grns': all_purchases.count(),
        'total_value': all_purchases.aggregate(total=Sum('total_amount'))['total'] or 0,
        'paid_amount': paid_total,
        'outstanding_amount': outstanding_total,
        'unpaid': unpaid_count,
        'pending_stock_update': all_purchases.filter(stock_updated=False, status='received').count(),
    }
    
    # Pagination
    from django.core.paginator import Paginator
    per_page = int(request.GET.get('per_page', 20))
    paginator = Paginator(purchases, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    companies = Company.objects.filter(is_active=True).order_by('company_name')
    
    context = {
        'purchases': page_obj,
        'companies': companies,
        'stats': stats,
        'page_obj': page_obj,
    }
    return render(request, 'products/purchase_list.html', context)


@login_required
def create_purchase(request):
    """Create new GRN"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get PO if specified
                po_id = request.POST.get('purchase_order')
                purchase_order = None
                if po_id:
                    purchase_order = PurchaseOrder.objects.get(pk=po_id)
                
                # Create GRN
                purchase = Purchase(
                    purchase_order=purchase_order,
                    company_id=request.POST.get('company'),
                    grn_date=timezone.now(),
                    invoice_date=request.POST.get('invoice_date') or None,
                    supplier_invoice_number=request.POST.get('supplier_invoice_number'),
                    status='draft',
                    notes=request.POST.get('notes', ''),
                    created_by=request.user,
                )
                
                purchase.save()  # Save to generate GRN number
                
                # Process items - new format using product IDs
                products = Product.objects.filter(is_active=True)
                items_created = 0
                
                for product in products:
                    # Get quantities for this product
                    packs = int(request.POST.get(f'packs_{product.id}', 0) or 0)
                    loose = int(request.POST.get(f'loose_{product.id}', 0) or 0)
                    foc_qty = int(request.POST.get(f'foc_{product.id}', 0) or 0)
                    
                    # Skip if no quantity entered (including FOC)
                    if packs == 0 and loose == 0 and foc_qty == 0:
                        continue
                    
                    # Get pricing from form
                    # price_{{id}} contains shop_price (price after shop discount from product table)
                    # discount_{{id}} contains company_discount_percentage (distributor discount to apply)
                    shop_price = request.POST.get(f'price_{product.id}', 0)
                    company_discount_pct = request.POST.get(f'discount_{product.id}', 0)
                    
                    # Convert to Decimal
                    shop_price = Decimal(shop_price) if shop_price else product.shop_price
                    company_discount_pct = Decimal(company_discount_pct) if company_discount_pct else product.company_discount_percentage
                    
                    # Create item with new field structure
                    # invoice_price = shop_price (from product table or form)
                    # PurchaseItem.save() will calculate: unit_price = shop_price - company_discount
                    PurchaseItem.objects.create(
                        purchase=purchase,
                        product=product,
                        packs=packs,
                        loose_bottles=loose,
                        bottles_per_pack=product.bottles_per_pack,
                        foc_quantity=foc_qty,
                        # Pricing fields
                        marked_price=product.marked_price,  # From product master
                        shop_discount_percentage=product.discount_percentage,  # From product master
                        invoice_price=shop_price,  # Shop price (marked - shop discount) from form or product
                        company_discount_percentage=company_discount_pct,  # Distributor discount from form
                        # unit_price, quantity, line_total will be auto-calculated in save()
                    )
                    items_created += 1
                
                if items_created == 0:
                    messages.warning(request, 'No items added to GRN. Please enter quantities for at least one product.')
                    purchase.delete()
                    companies = Company.objects.filter(is_active=True)
                    from products.utils import get_size_ordering
                    products = Product.objects.filter(is_active=True).select_related('company', 'category').annotate(
                        size_num=get_size_ordering()
                    ).order_by('size_num', 'marked_price', 'display_order', 'product_name')
                    purchase_orders = PurchaseOrder.objects.filter(status__in=['ordered', 'draft']).select_related('company').order_by('-order_date')
                    context = {
                        'companies': companies,
                        'products': products,
                        'purchase_orders': purchase_orders,
                    }
                    return render(request, 'products/create_purchase.html', context)
                
                # Calculate totals from items
                purchase.calculate_totals()
                
                messages.success(request, f'GRN {purchase.grn_number} created successfully!')
                return redirect('products:purchase_detail', pk=purchase.pk)
                
        except Exception as e:
            messages.error(request, f'Error creating GRN: {str(e)}')
    
    # GET request - check if PO specified in URL
    po_id = request.GET.get('po_id')
    selected_po = None
    if po_id:
        try:
            selected_po = PurchaseOrder.objects.select_related('company').get(pk=po_id)
        except PurchaseOrder.DoesNotExist:
            messages.warning(request, 'Purchase order not found.')
    
    companies = Company.objects.filter(is_active=True).order_by('company_name')
    from products.utils import get_size_ordering
    products = Product.objects.filter(is_active=True).select_related('company', 'category').annotate(
        size_num=get_size_ordering()
    ).order_by('size_num', 'marked_price', 'display_order', 'product_name')
    purchase_orders = PurchaseOrder.objects.filter(status__in=['ordered', 'draft']).select_related('company').order_by('-order_date')
    
    context = {
        'companies': companies,
        'products': products,
        'purchase_orders': purchase_orders,
        'selected_po': selected_po,
    }
    return render(request, 'products/create_purchase.html', context)


@login_required
def purchase_detail(request, pk):
    """GRN detail view"""
    purchase = get_object_or_404(Purchase.objects.select_related('company', 'created_by'), pk=pk)
    from products.utils import get_size_ordering
    items = purchase.items.select_related('product').annotate(
        size_num=get_size_ordering('product__size')
    ).order_by('size_num', 'product__marked_price', 'product__display_order', 'product__product_name')
    
    # Annotate each item with computed display values
    from decimal import Decimal
    for item in items:
        item.value_before_discount = item.quantity * item.invoice_price
        company_disc_amount = (item.value_before_discount * item.company_discount_percentage) / Decimal('100')
        item.company_discount_amount = company_disc_amount
    
    # Calculate summary
    summary = {
        'total_items': items.count(),
        'total_packs': sum(item.packs for item in items),
        'total_loose': sum(item.loose_bottles for item in items),
        'total_bottles': sum(item.quantity for item in items),
        'total_foc': sum(item.foc_quantity for item in items),
    }
    
    # Get returns settled with this GRN
    settled_returns = PurchaseReturn.objects.filter(replacement_grn=purchase).select_related('created_by')
    total_settled_amount = sum(ret.replacement_received_value or 0 for ret in settled_returns)
    
    context = {
        'purchase': purchase,
        'items': items,
        'summary': summary,
        'settled_returns': settled_returns,
        'total_settled_amount': total_settled_amount,
    }
    return render(request, 'products/purchase_detail.html', context)


@login_required
def edit_purchase(request, pk):
    """Edit a draft GRN — update header fields and items"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    purchase = get_object_or_404(Purchase.objects.select_related('company', 'purchase_order'), pk=pk)

    if purchase.status != 'draft':
        messages.error(request, 'Only draft GRNs can be edited.')
        return redirect('products:purchase_detail', pk=pk)

    if purchase.stock_updated:
        messages.error(request, 'Cannot edit a GRN that has already updated stock.')
        return redirect('products:purchase_detail', pk=pk)

    existing_items = purchase.items.select_related('product')
    # Map product_id → existing PurchaseItem for pre-fill
    existing_item_map = {item.product_id: item for item in existing_items}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Delete old items
                existing_items.delete()

                # Update header fields
                po_id = request.POST.get('purchase_order')
                if po_id:
                    purchase.purchase_order = PurchaseOrder.objects.get(pk=po_id)
                else:
                    purchase.purchase_order = None

                purchase.company_id = request.POST.get('company')
                purchase.invoice_date = request.POST.get('invoice_date') or None
                purchase.supplier_invoice_number = request.POST.get('supplier_invoice_number', '')
                purchase.notes = request.POST.get('notes', '')
                purchase.save()

                # Create new items
                products = Product.objects.filter(is_active=True)
                items_created = 0

                for product in products:
                    packs = int(request.POST.get(f'packs_{product.id}', 0) or 0)
                    loose = int(request.POST.get(f'loose_{product.id}', 0) or 0)
                    foc_qty = int(request.POST.get(f'foc_{product.id}', 0) or 0)

                    if packs == 0 and loose == 0 and foc_qty == 0:
                        continue

                    shop_price = request.POST.get(f'price_{product.id}', 0)
                    company_discount_pct = request.POST.get(f'discount_{product.id}', 0)
                    shop_price = Decimal(shop_price) if shop_price else product.shop_price
                    company_discount_pct = Decimal(company_discount_pct) if company_discount_pct else product.company_discount_percentage

                    PurchaseItem.objects.create(
                        purchase=purchase,
                        product=product,
                        packs=packs,
                        loose_bottles=loose,
                        bottles_per_pack=product.bottles_per_pack,
                        foc_quantity=foc_qty,
                        marked_price=product.marked_price,
                        shop_discount_percentage=product.discount_percentage,
                        invoice_price=shop_price,
                        company_discount_percentage=company_discount_pct,
                    )
                    items_created += 1

                if items_created == 0:
                    raise ValueError('Please add at least one product with quantities.')

                purchase.calculate_totals()
                messages.success(request, f'GRN {purchase.grn_number} updated successfully!')
                return redirect('products:purchase_detail', pk=purchase.pk)

        except Exception as e:
            messages.error(request, f'Error updating GRN: {str(e)}')

    companies = Company.objects.filter(is_active=True).order_by('company_name')
    from products.utils import get_size_ordering
    products = Product.objects.filter(is_active=True).select_related('company', 'category').annotate(
        size_num=get_size_ordering()
    ).order_by('size_num', 'marked_price', 'display_order', 'product_name')
    purchase_orders = PurchaseOrder.objects.filter(
        status__in=['ordered', 'draft']
    ).select_related('company').order_by('-order_date')

    # Build JSON-safe dict from existing items for JS pre-fill
    import json
    existing_items_json = json.dumps({
        str(pid): {
            'packs': item.packs,
            'loose': item.loose_bottles,
            'foc': item.foc_quantity,
            'price': float(item.invoice_price),
            'discount': float(item.company_discount_percentage),
        }
        for pid, item in existing_item_map.items()
    })

    context = {
        'companies': companies,
        'products': products,
        'purchase_orders': purchase_orders,
        'purchase': purchase,
        'existing_items_json': existing_items_json,
        'is_edit': True,
    }
    return render(request, 'products/edit_purchase.html', context)


@login_required
def delete_purchase(request, pk):
    """Delete a draft GRN (only draft status allowed)"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    purchase = get_object_or_404(Purchase, pk=pk)
    
    if purchase.status != 'draft':
        messages.error(request, 'Only draft GRNs can be deleted.')
        return redirect('products:purchase_detail', pk=pk)
    
    if purchase.stock_updated:
        messages.error(request, 'Cannot delete a GRN that has already updated stock.')
        return redirect('products:purchase_detail', pk=pk)
    
    if request.method == 'POST':
        grn_number = purchase.grn_number
        purchase.items.all().delete()
        purchase.delete()
        messages.success(request, f'GRN {grn_number} has been deleted.')
        return redirect('products:purchase_list')
    
    return redirect('products:purchase_detail', pk=pk)


@login_required
def update_purchase_stock(request, pk):
    """Update stock from GRN (receive goods)"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    purchase = get_object_or_404(Purchase, pk=pk)
    
    # ENHANCED CHECK: Verify stock not already updated
    # Check both flag AND stock movements (prevents race condition)
    if purchase.stock_updated or StockMovement.objects.filter(
        reference_number=purchase.grn_number,
        movement_type='purchase'
    ).exists():
        messages.warning(
            request, 
            f'Stock has already been updated for {purchase.grn_number}. '
            f'Cannot update again to prevent duplicate entries.'
        )
        return redirect('products:purchase_detail', pk=pk)
    
    # VALIDATION: Ensure GRN has items
    items = purchase.items.all()
    if not items.exists():
        messages.error(request, f'Cannot update stock: {purchase.grn_number} has no items.')
        return redirect('products:purchase_detail', pk=pk)
    
    # VALIDATION: Check all items have valid quantities (regular OR FOC)
    invalid_items = items.filter(
        Q(quantity__lte=0, foc_quantity__lte=0) | Q(packs=0, loose_bottles=0, foc_quantity__lte=0)
    )
    if invalid_items.exists():
        messages.error(
            request,
            f'Cannot update stock: {invalid_items.count()} item(s) have zero or negative quantities. '
            f'Please edit GRN and add quantities before receiving.'
        )
        return redirect('products:purchase_detail', pk=pk)
    
    try:
        with transaction.atomic():
            # SET FLAG FIRST (race condition prevention)
            # If transaction fails, flag rolls back with stock changes
            purchase.stock_updated = True
            purchase.status = 'received'
            purchase.received_by = request.user
            purchase.save()

            # Auto-mark the linked PO as received
            if purchase.purchase_order and purchase.purchase_order.status == 'ordered':
                purchase.purchase_order.status = 'received'
                purchase.purchase_order.received_date = purchase.grn_date.date() if hasattr(purchase.grn_date, 'date') else purchase.grn_date
                purchase.purchase_order.save(update_fields=['status', 'received_date'])
            
            # Track totals for confirmation message
            items_updated = 0
            total_bottles = 0
            total_foc = 0
            
            for item in items:
                # Total received = quantity + FOC (both added to main stock)
                total_received = item.quantity + item.foc_quantity
                
                # VALIDATION: Prevent negative total (edge case check)
                if total_received <= 0:
                    raise ValueError(
                        f'Invalid quantity for {item.product.product_name}: '
                        f'{item.quantity} bottles + {item.foc_quantity} FOC = {total_received} total'
                    )
                
                # Capture stock before update
                previous_qty = item.product.quantity_in_stock
                
                # Update product stock
                item.product.quantity_in_stock += total_received
                item.product.save()
                
                # Calculate effective cost per unit (spread across qty + FOC)
                effective_unit_cost = (item.unit_price * item.quantity) / total_received if total_received > 0 else item.unit_price
                
                # Create stock movement with comprehensive details
                StockMovement.objects.create(
                    product=item.product,
                    movement_type='purchase',
                    quantity=total_received,
                    previous_quantity=previous_qty,
                    new_quantity=item.product.quantity_in_stock,
                    reference_number=purchase.grn_number,
                    notes=f'GRN: {purchase.grn_number} - Qty: {item.quantity}, FOC: {item.foc_quantity}',
                    created_by=request.user,
                    unit_cost=effective_unit_cost,
                    total_cost=effective_unit_cost * total_received,
                )
                
                # Create FIFO cost layer
                from products.models import FIFOCostLayer
                FIFOCostLayer.create_layer(
                    product=item.product,
                    qty=total_received,
                    unit_cost=effective_unit_cost,
                    source='purchase',
                    reference=purchase.grn_number,
                )
                
                items_updated += 1
                total_bottles += item.quantity
                total_foc += item.foc_quantity
                
                # Create FOC value transaction if FOC received
                if item.foc_quantity > 0:
                    # Get or create FOC account for this company
                    from products.models import FOCValueAccount, FOCValueTransaction
                    foc_account, created = FOCValueAccount.objects.get_or_create(
                        company=purchase.company,
                        defaults={'created_by': request.user}
                    )
                    
                    # Create FOC received transaction
                    FOCValueTransaction.objects.create(
                        foc_account=foc_account,
                        transaction_type='foc_received',
                        transaction_date=purchase.grn_date,
                        product=item.product,
                        foc_quantity=item.foc_quantity,
                        shop_price_at_time=item.product.shop_price,
                        reference_number=purchase.grn_number,
                        purchase_item=item,
                        notes=f'FOC received from {purchase.company.company_name}',
                        created_by=request.user
                    )
            
            # Enhanced success message with details
            messages.success(
                request, 
                f'✓ Stock updated successfully for {purchase.grn_number}! '
                f'{items_updated} product(s) received: '
                f'{total_bottles:,} bottles + {total_foc:,} FOC = {total_bottles + total_foc:,} total added to inventory.'
            )
            
    except Exception as e:
        # If any error occurs, transaction rolls back (including stock_updated flag)
        messages.error(request, f'Error updating stock: {str(e)}')
    
    return redirect('products:purchase_detail', pk=pk)


# Purchase Returns

@login_required
def purchase_return_list(request):
    """List all purchase returns with advanced filtering"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    # Check for export request
    if request.GET.get('export') == 'excel':
        return export_returns_excel(request)
    
    returns = PurchaseReturn.objects.select_related('company', 'created_by', 'approved_by').all()
    
    # Search functionality
    search = request.GET.get('search', '').strip()
    if search:
        returns = returns.filter(
            Q(pr_number__icontains=search) |
            Q(company__company_name__icontains=search) |
            Q(detailed_reason__icontains=search)
        )
    
    # Date range filter
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date:
        returns = returns.filter(return_date__date__gte=from_date)
    if to_date:
        returns = returns.filter(return_date__date__lte=to_date)
    
    # Filters
    return_type = request.GET.get('return_type')
    if return_type:
        returns = returns.filter(return_type=return_type)
    
    status = request.GET.get('status')
    if status:
        returns = returns.filter(status=status)
    
    company_id = request.GET.get('company')
    if company_id:
        returns = returns.filter(company_id=company_id)
    
    settlement_type = request.GET.get('settlement_type')
    if settlement_type:
        returns = returns.filter(settlement_type=settlement_type)
    
    # Sorting
    sort = request.GET.get('sort', '-return_date')
    returns = returns.order_by(sort)
    
    # Get all returns for comprehensive stats (before pagination)
    all_returns = PurchaseReturn.objects.all()
    stats = {
        'total_returns': all_returns.count(),
        'expired_count': all_returns.filter(return_type='expired').count(),
        'damaged_count': all_returns.filter(return_type='damaged').count(),
        'pending': all_returns.filter(status='pending').count(),
        'settled_count': all_returns.filter(status='settled').count(),
        'total_value': all_returns.aggregate(total=Sum('total_amount'))['total'] or 0,
    }
    
    # Pagination
    from django.core.paginator import Paginator
    per_page = int(request.GET.get('per_page', 20))
    paginator = Paginator(returns, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    companies = Company.objects.filter(is_active=True).order_by('company_name')
    
    context = {
        'returns': page_obj,
        'companies': companies,
        'stats': stats,
        'page_obj': page_obj,
    }
    return render(request, 'products/purchase_return_list.html', context)


@login_required
def create_purchase_return(request):
    """Create purchase return"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                return_type = request.POST.get('return_type', 'damaged')
                purchase_return = PurchaseReturn(
                    company_id=request.POST.get('company'),
                    return_date=timezone.now(),
                    return_type=return_type,
                    return_reason=return_type,  # Use return_type as reason for simplicity
                    detailed_reason=request.POST.get('detailed_reason'),
                    settlement_type='refund',  # Default: Cash refund or replacement
                    created_by=request.user,
                )
                
                purchase_return.save()
                
                # Process items - new format uses quantity_<product_id>
                total = Decimal('0')
                items_created = 0
                
                for key, value in request.POST.items():
                    if key.startswith('quantity_') and value and int(value) > 0:
                        product_id = key.replace('quantity_', '')
                        product = Product.objects.get(pk=product_id)
                        qty = int(value)
                        
                        cost_per_unit = product.cost_after_foc if product.cost_after_foc else Decimal('0')

                        # Auto-transfer from resaleable if qty exceeds non-resaleable
                        excess = qty - product.non_resaleable_stock
                        if excess > 0:
                            total_available = product.non_resaleable_stock + product.quantity_in_stock
                            if total_available < qty:
                                raise ValueError(
                                    f'Cannot return {qty} {product.product_name}: '
                                    f'Only {total_available} total available '
                                    f'({product.non_resaleable_stock} non-resaleable + {product.quantity_in_stock} resaleable)'
                                )
                            prev_res = product.quantity_in_stock
                            prev_non_res = product.non_resaleable_stock
                            product.quantity_in_stock -= excess
                            product.non_resaleable_stock += excess
                            product.save()
                            StockMovement.objects.create(
                                product=product, movement_type='non_resaleable_in', stock_type='resaleable',
                                quantity=-excess, previous_quantity=prev_res, new_quantity=product.quantity_in_stock,
                                reference_number=purchase_return.pr_number,
                                notes=f'Auto-transferred to non-resaleable for Purchase Return: {purchase_return.pr_number}',
                                created_by=request.user, unit_cost=cost_per_unit, total_cost=cost_per_unit * excess,
                            )
                            StockMovement.objects.create(
                                product=product, movement_type='non_resaleable_in', stock_type='non_resaleable',
                                quantity=excess, previous_quantity=prev_non_res, new_quantity=product.non_resaleable_stock,
                                reference_number=purchase_return.pr_number,
                                notes=f'Auto-transferred from resaleable for Purchase Return: {purchase_return.pr_number}',
                                created_by=request.user, unit_cost=cost_per_unit, total_cost=cost_per_unit * excess,
                            )

                        # Reduce non-resaleable stock immediately
                        previous_stock = product.non_resaleable_stock
                        product.non_resaleable_stock -= qty
                        product.save()

                        # Create stock movement record
                        StockMovement.objects.create(
                            product=product,
                            movement_type='return_to_company',
                            quantity=-qty,  # Negative because removing from stock
                            previous_quantity=previous_stock,
                            new_quantity=product.non_resaleable_stock,
                            reference_number=purchase_return.pr_number,
                            notes=f'Purchase Return: {purchase_return.pr_number} - Removed from non-resaleable stock',
                            created_by=request.user,
                            unit_cost=cost_per_unit,
                            total_cost=cost_per_unit * qty,
                        )
                        
                        # Create item with pricing fields - save() will auto-calculate unit_price and line_total
                        submitted_marked_price = request.POST.get(f'marked_price_{product_id}')
                        marked_price = Decimal(submitted_marked_price) if submitted_marked_price else product.marked_price
                        shop_discount = product.discount_percentage
                        invoice_price = marked_price - (marked_price * shop_discount / 100)
                        PurchaseReturnItem.objects.create(
                            purchase_return=purchase_return,
                            product=product,
                            quantity=qty,
                            marked_price=marked_price,
                            shop_discount_percentage=shop_discount,
                            invoice_price=invoice_price,
                            company_discount_percentage=product.company_discount_percentage,
                            # unit_price and line_total will be auto-calculated in save()
                        )
                        items_created += 1
                
                if items_created == 0:
                    messages.error(request, 'Please select at least one product with quantity.')
                    purchase_return.delete()
                    raise ValueError('No products selected')
                
                # Use calculate_totals method
                purchase_return.calculate_totals()
                
                messages.success(request, f'Purchase return {purchase_return.pr_number} created successfully!')
                return redirect('products:purchase_return_detail', pk=purchase_return.pk)
                
        except Exception as e:
            messages.error(request, f'Error creating purchase return: {str(e)}')
    
    companies = Company.objects.filter(is_active=True)
    from products.utils import get_size_ordering
    products = Product.objects.filter(is_active=True).annotate(
        size_num=get_size_ordering()
    ).order_by('size_num', 'marked_price', 'display_order', 'product_name')
    
    context = {
        'companies': companies,
        'products': products,
    }
    return render(request, 'products/create_purchase_return.html', context)


@login_required
def purchase_return_detail(request, pk):
    """Purchase return detail"""
    purchase_return = get_object_or_404(
        PurchaseReturn.objects.select_related('company', 'created_by', 'approved_by', 'purchase', 'replacement_grn'), 
        pk=pk
    )
    from products.utils import get_size_ordering
    items = purchase_return.items.select_related('product').annotate(
        size_num=get_size_ordering('product__size')
    ).order_by('size_num', 'product__marked_price', 'product__display_order', 'product__product_name')
    
    # Calculate summary
    summary = {
        'total_qty': sum(item.quantity for item in items),
    }
    
    # Get settlement records
    settlements = purchase_return.settlements.all().select_related('replacement_grn')
    
    # Calculate settlement percentage - ensure it's a float/int
    percentage = float(purchase_return.settlement_percentage) if purchase_return.settlement_percentage else 0
    
    # Calculate remaining amount to settle
    approved_amount = purchase_return.approved_amount or purchase_return.total_amount
    already_settled = purchase_return.total_settled_amount
    remaining_to_settle = approved_amount - already_settled
    
    # Get GRNs from same company for replacement linking (used in modal)
    # Only show GRNs not linked to POs (direct purchases/replacements)
    available_grns = Purchase.objects.filter(
        company=purchase_return.company,
        purchase_order__isnull=True  # Only GRNs without PO link
    ).order_by('-grn_date')[:50]  # Limit to recent 50 GRNs
    
    context = {
        'purchase_return': purchase_return,
        'items': items,
        'summary': summary,
        'settlements': settlements,
        'percentage': percentage,
        'remaining_to_settle': remaining_to_settle,
        'available_grns': available_grns,
    }
    return render(request, 'products/purchase_return_detail.html', context)


@login_required
def edit_purchase_return(request, pk):
    """Edit a pending purchase return — update items, quantities, reason, type"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')

    purchase_return = get_object_or_404(
        PurchaseReturn.objects.select_related('company'), pk=pk
    )

    if purchase_return.status != 'pending':
        messages.warning(request, 'Only pending returns can be edited.')
        return redirect('products:purchase_return_detail', pk=pk)

    existing_items = purchase_return.items.select_related('product')
    # Map product_id → existing item for pre-fill
    existing_qty_map = {item.product_id: item for item in existing_items}

    if request.method == 'POST':
        try:
            with transaction.atomic():
                # --- 1. Reverse old non_resaleable_stock deductions ---
                for item in existing_items:
                    product = item.product
                    product.non_resaleable_stock += item.quantity
                    product.save()

                # Delete old stock movements for this PR
                StockMovement.objects.filter(
                    reference_number=purchase_return.pr_number,
                    movement_type='return_to_company',
                ).delete()

                # Delete old items
                existing_items.delete()

                # --- 2. Update header fields ---
                new_company_id = request.POST.get('company')
                purchase_return.company_id = new_company_id
                return_type = request.POST.get('return_type', 'damaged')
                purchase_return.return_type = return_type
                purchase_return.return_reason = return_type
                purchase_return.detailed_reason = request.POST.get('detailed_reason', '')
                purchase_return.save()

                # --- 3. Create new items ---
                items_created = 0
                for key, value in request.POST.items():
                    if key.startswith('quantity_') and value and int(value) > 0:
                        product_id = key.replace('quantity_', '')
                        product = Product.objects.get(pk=product_id)
                        qty = int(value)

                        cost_per_unit = product.cost_after_foc if product.cost_after_foc else Decimal('0')

                        # Auto-transfer from resaleable if qty exceeds non-resaleable
                        excess = qty - product.non_resaleable_stock
                        if excess > 0:
                            total_available = product.non_resaleable_stock + product.quantity_in_stock
                            if total_available < qty:
                                raise ValueError(
                                    f'Cannot return {qty} {product.product_name}: '
                                    f'Only {total_available} total available '
                                    f'({product.non_resaleable_stock} non-resaleable + {product.quantity_in_stock} resaleable)'
                                )
                            prev_res = product.quantity_in_stock
                            prev_non_res = product.non_resaleable_stock
                            product.quantity_in_stock -= excess
                            product.non_resaleable_stock += excess
                            product.save()
                            StockMovement.objects.create(
                                product=product, movement_type='non_resaleable_in', stock_type='resaleable',
                                quantity=-excess, previous_quantity=prev_res, new_quantity=product.quantity_in_stock,
                                reference_number=purchase_return.pr_number,
                                notes=f'Auto-transferred to non-resaleable for Purchase Return: {purchase_return.pr_number}',
                                created_by=request.user, unit_cost=cost_per_unit, total_cost=cost_per_unit * excess,
                            )
                            StockMovement.objects.create(
                                product=product, movement_type='non_resaleable_in', stock_type='non_resaleable',
                                quantity=excess, previous_quantity=prev_non_res, new_quantity=product.non_resaleable_stock,
                                reference_number=purchase_return.pr_number,
                                notes=f'Auto-transferred from resaleable for Purchase Return: {purchase_return.pr_number}',
                                created_by=request.user, unit_cost=cost_per_unit, total_cost=cost_per_unit * excess,
                            )

                        # Reduce non-resaleable stock
                        previous_stock = product.non_resaleable_stock
                        product.non_resaleable_stock -= qty
                        product.save()

                        # Stock movement
                        StockMovement.objects.create(
                            product=product,
                            movement_type='return_to_company',
                            quantity=-qty,
                            previous_quantity=previous_stock,
                            new_quantity=product.non_resaleable_stock,
                            reference_number=purchase_return.pr_number,
                            notes=f'Purchase Return (Edit): {purchase_return.pr_number} - Removed from non-resaleable stock',
                            created_by=request.user,
                            unit_cost=cost_per_unit,
                            total_cost=cost_per_unit * qty,
                        )

                        # Get price/discount overrides from form
                        marked_price_input = request.POST.get(f'marked_price_{product_id}')
                        shop_disc_input = request.POST.get(f'shop_discount_{product_id}')
                        company_disc_input = request.POST.get(f'company_discount_{product_id}')
                        marked_price = Decimal(marked_price_input) if marked_price_input else product.marked_price
                        shop_disc = Decimal(shop_disc_input) if shop_disc_input else product.discount_percentage
                        company_disc = Decimal(company_disc_input) if company_disc_input else product.company_discount_percentage
                        invoice_price = marked_price * (1 - shop_disc / 100)

                        PurchaseReturnItem.objects.create(
                            purchase_return=purchase_return,
                            product=product,
                            quantity=qty,
                            marked_price=marked_price,
                            shop_discount_percentage=shop_disc,
                            invoice_price=invoice_price,
                            company_discount_percentage=company_disc,
                        )
                        items_created += 1

                if items_created == 0:
                    raise ValueError('Please select at least one product with quantity.')

                purchase_return.calculate_totals()
                messages.success(request, f'Purchase return {purchase_return.pr_number} updated successfully!')
                return redirect('products:purchase_return_detail', pk=purchase_return.pk)

        except Exception as e:
            messages.error(request, f'Error updating purchase return: {str(e)}')

    companies = Company.objects.filter(is_active=True)
    from products.utils import get_size_ordering
    products = Product.objects.filter(is_active=True).annotate(
        size_num=get_size_ordering()
    ).order_by('size_num', 'marked_price', 'display_order', 'product_name')

    # Build JSON-safe dict from existing items for JS pre-fill
    import json
    existing_items_json = json.dumps({
        str(pid): {
            'quantity': item.quantity,
            'marked_price': float(item.marked_price),
            'shop_discount': float(item.shop_discount_percentage),
            'company_discount': float(item.company_discount_percentage),
        }
        for pid, item in existing_qty_map.items()
    })

    context = {
        'companies': companies,
        'products': products,
        'purchase_return': purchase_return,
        'existing_items_json': existing_items_json,
        'is_edit': True,
    }
    return render(request, 'products/edit_purchase_return.html', context)


@login_required
@login_required
def approve_purchase_return(request, pk):
    """Approve purchase return and reduce stock - Stage 1: Send to Supplier"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    purchase_return = get_object_or_404(PurchaseReturn, pk=pk)
    
    if purchase_return.status != 'pending':
        messages.warning(request, 'This return has already been processed.')
        return redirect('products:purchase_return_detail', pk=pk)
    
    try:
        with transaction.atomic():
            from products.models import FIFOCostLayer
            for item in purchase_return.items.all():
                # non_resaleable_stock was already reduced when the return was created.
                # Only compute cost for the status-change movement record.
                cost_per_unit = item.product.cost_after_foc if item.product.cost_after_foc else Decimal('0')

                # Create stock movement to record the approval event (no stock change)
                StockMovement.objects.create(
                    product=item.product,
                    movement_type='purchase_return',
                    quantity=0,
                    previous_quantity=item.product.non_resaleable_stock,
                    new_quantity=item.product.non_resaleable_stock,
                    reference_number=purchase_return.pr_number,
                    notes=f'Purchase Return Sent to Supplier: {purchase_return.pr_number}',
                    created_by=request.user,
                    unit_cost=cost_per_unit,
                    total_cost=cost_per_unit * item.quantity,
                )
            
            # New workflow: Change to sent_to_supplier instead of approved
            purchase_return.status = 'sent_to_supplier'
            purchase_return.stock_updated = True
            purchase_return.sent_by = request.user
            purchase_return.sent_at = timezone.now()
            purchase_return.sent_date = timezone.localdate()
            purchase_return.save()
            
            messages.success(request, f'Purchase return {purchase_return.pr_number} sent to supplier! Stock has been reduced. Waiting for company approval.')
            
    except Exception as e:
        messages.error(request, f'Error processing return: {str(e)}')
    
    return redirect('products:purchase_return_detail', pk=pk)


@login_required
def record_company_approval(request, pk):
    """Record company approval with approved amount - Stage 2: Company Response"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    purchase_return = get_object_or_404(PurchaseReturn, pk=pk)
    
    # Accept both new 'sent_to_supplier' and old 'sent' status for backward compatibility
    if purchase_return.status not in ['sent_to_supplier', 'sent']:
        messages.warning(request, 'Return must be sent to supplier first.')
        return redirect('products:purchase_return_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get approved amount from company
                approved_amount_str = request.POST.get('approved_amount', '0')
                approved_amount = Decimal(approved_amount_str) if approved_amount_str else Decimal('0')
                
                # Get company approval date
                company_approved_date = request.POST.get('company_approved_date')
                
                # Update purchase return
                purchase_return.approved_amount = approved_amount
                purchase_return.company_approved_date = company_approved_date
                purchase_return.status = 'company_approved'
                purchase_return.approved_by = request.user
                purchase_return.approved_at = timezone.now()
                purchase_return.save()
                
                # Auto-create company transaction for the approved amount
                purchase_return.create_return_transaction()
                
                messages.success(request, f'Company approval recorded! Approved amount: Rs. {approved_amount:,.2f}. You can now settle this return.')
                
        except Exception as e:
            messages.error(request, f'Error recording approval: {str(e)}')
    
    return redirect('products:purchase_return_detail', pk=pk)


@login_required
def mark_return_sent(request, pk):
    """Mark purchase return as sent to company"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    purchase_return = get_object_or_404(PurchaseReturn, pk=pk)
    
    if purchase_return.status != 'approved':
        messages.warning(request, 'Return must be approved first.')
        return redirect('products:purchase_return_detail', pk=pk)
    
    purchase_return.status = 'sent'
    purchase_return.sent_date = timezone.localdate()
    purchase_return.save()
    
    messages.success(request, f'Purchase return {purchase_return.pr_number} marked as sent to company.')
    return redirect('products:purchase_return_detail', pk=pk)


@login_required
def update_return_settlement(request, pk):
    """Update settlement details with multiple settlement methods (POST only)"""
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    purchase_return = get_object_or_404(PurchaseReturn, pk=pk)
    
    # Check if already company approved
    if purchase_return.status != 'company_approved':
        messages.warning(request, 'Return must be company approved first.')
        return redirect('products:purchase_return_detail', pk=pk)
    
    # Redirect to detail page if not POST (this is an action endpoint, not a form page)
    if request.method != 'POST':
        return redirect('products:purchase_return_detail', pk=pk)
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Get settlement methods, amounts, and references as arrays
                methods = request.POST.getlist('settlement_methods[]')
                amounts = request.POST.getlist('settlement_amounts[]')
                references = request.POST.getlist('settlement_references[]')
                
                if not methods or not amounts:
                    messages.error(request, 'No settlement methods provided.')
                    return redirect('products:purchase_return_detail', pk=pk)
                
                # Calculate total settlement
                from decimal import Decimal as D
                new_settlement_total = sum([D(amt) for amt in amounts if amt])
                approved_amount = purchase_return.approved_amount or D('0')
                
                # Get existing settled amount (don't delete - we're adding to it)
                existing_settled = purchase_return.total_settled_amount
                
                # Calculate new total after adding these settlements
                new_total_settled = existing_settled + new_settlement_total
                
                # Validate total doesn't exceed approved amount
                if new_total_settled > approved_amount:
                    remaining = approved_amount - existing_settled
                    messages.error(request, 
                        f'Cannot add Rs. {new_settlement_total:,.2f} in new settlements. '
                        f'Already settled: Rs. {existing_settled:,.2f} | '
                        f'Remaining available: Rs. {remaining:,.2f} | '
                        f'Approved amount: Rs. {approved_amount:,.2f}')
                    return redirect('products:purchase_return_detail', pk=pk)
                
                # Validate we're actually adding something
                if new_settlement_total <= 0:
                    messages.error(request, 'Please enter settlement amounts.')
                    return redirect('products:purchase_return_detail', pk=pk)
                
                # Process each settlement method
                settlement_summary = []
                primary_method = methods[0] if methods else 'credit_note'
                total_replacement_value = D('0')
                used_grn_ids = set()  # Track GRNs to prevent duplicates
                
                # Get existing GRN IDs to prevent duplicate usage
                existing_grn_ids = set(
                    purchase_return.settlements.filter(
                        settlement_method='replacement',
                        replacement_grn__isnull=False
                    ).values_list('replacement_grn_id', flat=True)
                )
                used_grn_ids.update(existing_grn_ids)
                
                # DO NOT clear existing settlements - we're adding to them
                
                for i, method in enumerate(methods):
                    amount = D(amounts[i]) if i < len(amounts) else D('0')
                    reference = references[i] if i < len(references) else ''
                    
                    if method == 'replacement' and reference:
                        try:
                            replacement_grn = Purchase.objects.get(pk=reference)
                            
                            # Validate GRN is not PO-based
                            if replacement_grn.purchase_order:
                                raise ValueError(
                                    f'{replacement_grn.grn_number} is linked to Purchase Order '
                                    f'{replacement_grn.purchase_order.po_number}. Cannot use PO-based GRNs for settlements.'
                                )
                            
                            # Validate GRN not used multiple times in same settlement
                            if replacement_grn.pk in used_grn_ids:
                                raise ValueError(
                                    f'{replacement_grn.grn_number} is already used in this settlement. '
                                    f'Each GRN can only be used once per return settlement.'
                                )
                            
                            # Get current outstanding (real-time check)
                            current_outstanding = replacement_grn.amount_outstanding
                            
                            # CRITICAL: Prevent negative outstanding (over-settlement)
                            if current_outstanding <= 0:
                                raise ValueError(
                                    f'{replacement_grn.grn_number} is already fully settled/paid. '
                                    f'Cannot use for additional settlements. '
                                    f'Outstanding balance: Rs. {current_outstanding:,.2f}'
                                )
                            
                            # Validate settlement doesn't exceed GRN outstanding balance
                            if amount > current_outstanding:
                                raise ValueError(
                                    f'Cannot settle Rs. {amount:,.2f} using {replacement_grn.grn_number}. '
                                    f'GRN outstanding balance is only Rs. {current_outstanding:,.2f} '
                                    f'(Total: Rs. {replacement_grn.total_amount:,.2f}, '
                                    f'Already Paid: Rs. {replacement_grn.total_paid:,.2f}, '
                                    f'Settled via Returns: Rs. {replacement_grn.total_settled_via_returns:,.2f})'
                                )
                            
                            # Warning: Near full settlement (within Rs. 10)
                            remaining_after = current_outstanding - amount
                            if remaining_after < D('10') and remaining_after > 0:
                                messages.warning(
                                    request,
                                    f'Note: {replacement_grn.grn_number} will have only Rs. {remaining_after:.2f} '
                                    f'outstanding after this settlement. Consider allocating full amount.'
                                )
                            
                            # Mark GRN as used
                            used_grn_ids.add(replacement_grn.pk)
                            
                            # Create settlement record
                            PurchaseReturnSettlement.objects.create(
                                purchase_return=purchase_return,
                                settlement_method='replacement',
                                settlement_amount=amount,
                                replacement_grn=replacement_grn,
                                created_by=request.user
                            )
                            
                            # Update legacy fields for backward compatibility (use first/primary GRN)
                            if not purchase_return.replacement_grn:
                                purchase_return.replacement_grn = replacement_grn
                            
                            total_replacement_value += amount
                            settlement_summary.append(f'Replacement GRN {replacement_grn.grn_number}: Rs. {amount:.2f}')
                        except (Purchase.DoesNotExist, ValueError):
                            messages.warning(request, f'GRN {reference} not found, skipped.')
                    
                    elif method == 'refund':
                        # Create settlement record with cash audit trail
                        from django.utils import timezone
                        settlement = PurchaseReturnSettlement.objects.create(
                            purchase_return=purchase_return,
                            settlement_method='refund',
                            settlement_amount=amount,
                            refund_reference=reference if reference else 'Cash refund',
                            cash_received_date=timezone.localdate(),  # Auto-set to today
                            cash_receipt_number=reference if reference else None,
                            cash_verified_by=request.user,  # User recording the settlement
                            created_by=request.user
                        )
                        
                        # CRITICAL: Create CompanyTransaction for cash receipt
                        # This settles the receivable created when return was approved
                        account, created = CompanyAccount.objects.get_or_create(
                            company=purchase_return.company,
                            defaults={
                                'opening_balance': D('0'),
                                'opening_date': timezone.localdate(),
                                'created_by': request.user
                            }
                        )
                        
                        CompanyTransaction.objects.create(
                            company_account=account,
                            transaction_type='settlement',
                            transaction_date=timezone.localdate(),
                            reference_number=f'{purchase_return.pr_number}-REFUND',
                            amount=amount,
                            settlement_method='cash',
                            notes=f'Cash refund received for {purchase_return.pr_number}',
                            created_by=request.user
                        )
                        
                        messages.success(request, f'Cash refund of Rs. {amount:,.2f} recorded successfully!')
                        settlement_summary.append(f'Cash refund: Rs. {amount:.2f}')
                
                # Update settlement_status based on total settled
                approved_amount = purchase_return.approved_amount or D('0')
                total_settled = purchase_return.total_settled_amount  # This auto-calculates from settlements
                
                if total_settled >= approved_amount:
                    purchase_return.settlement_status = 'fully_settled'
                    purchase_return.status = 'settled'  # Only mark as settled when 100% complete
                elif total_settled > 0:
                    purchase_return.settlement_status = 'partial'
                    purchase_return.status = 'company_approved'  # Keep as approved for partial settlements
                else:
                    purchase_return.settlement_status = 'pending'
                    purchase_return.status = 'company_approved'  # Keep as approved if no settlement yet
                
                purchase_return.save()
                
                messages.success(request, 'Settlement updated successfully!')
                return redirect('products:purchase_return_detail', pk=pk)
                
        except Exception as e:
            messages.error(request, f'Error updating settlement: {str(e)}')
            return redirect('products:purchase_return_detail', pk=pk)


# Export Functions
def export_purchases_excel(request):
    """Export purchases to Excel"""
    purchases = Purchase.objects.select_related('company', 'created_by').all()
    
    # Apply same filters as list view
    search = request.GET.get('search', '').strip()
    if search:
        purchases = purchases.filter(
            Q(grn_number__icontains=search) |
            Q(supplier_invoice_number__icontains=search) |
            Q(company__company_name__icontains=search)
        )
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date:
        purchases = purchases.filter(grn_date__date__gte=from_date)
    if to_date:
        purchases = purchases.filter(grn_date__date__lte=to_date)
    
    status = request.GET.get('status')
    if status:
        purchases = purchases.filter(status=status)
    
    company_id = request.GET.get('company')
    if company_id:
        purchases = purchases.filter(company_id=company_id)
    
    payment_status = request.GET.get('payment_status')
    if payment_status:
        purchases = purchases.filter(payment_status=payment_status)
    
    sort = request.GET.get('sort', '-grn_date')
    purchases = purchases.order_by(sort)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchases"
    
    # Header style
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['GRN Number', 'Company', 'GRN Date', 'Invoice No.', 'Total Amount', 
               'Outstanding', 'Status', 'Payment Status', 'Stock Updated', 'Created By']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, purchase in enumerate(purchases, 2):
        ws.cell(row=row, column=1, value=purchase.grn_number)
        ws.cell(row=row, column=2, value=purchase.company.company_name)
        ws.cell(row=row, column=3, value=purchase.grn_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=4, value=purchase.supplier_invoice_number or '-')
        ws.cell(row=row, column=5, value=float(purchase.total_amount))
        ws.cell(row=row, column=6, value=float(purchase.outstanding_amount))
        ws.cell(row=row, column=7, value=purchase.get_status_display())
        ws.cell(row=row, column=8, value=purchase.get_payment_status_display())
        ws.cell(row=row, column=9, value='Yes' if purchase.stock_updated else 'No')
        ws.cell(row=row, column=10, value=purchase.created_by.get_full_name())
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=purchases_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


def export_returns_excel(request):
    """Export purchase returns to Excel"""
    returns = PurchaseReturn.objects.select_related('company', 'created_by', 'approved_by').all()
    
    # Apply same filters as list view
    search = request.GET.get('search', '').strip()
    if search:
        returns = returns.filter(
            Q(pr_number__icontains=search) |
            Q(company__company_name__icontains=search) |
            Q(detailed_reason__icontains=search)
        )
    
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    if from_date:
        returns = returns.filter(return_date__date__gte=from_date)
    if to_date:
        returns = returns.filter(return_date__date__lte=to_date)
    
    return_type = request.GET.get('return_type')
    if return_type:
        returns = returns.filter(return_type=return_type)
    
    status = request.GET.get('status')
    if status:
        returns = returns.filter(status=status)
    
    company_id = request.GET.get('company')
    if company_id:
        returns = returns.filter(company_id=company_id)
    
    settlement_type = request.GET.get('settlement_type')
    if settlement_type:
        returns = returns.filter(settlement_type=settlement_type)
    
    sort = request.GET.get('sort', '-return_date')
    returns = returns.order_by(sort)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Purchase Returns"
    
    # Header style
    header_fill = PatternFill(start_color="f093fb", end_color="f093fb", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = ['PR Number', 'Company', 'Return Date', 'Return Type', 'Reason', 
               'Total Amount', 'Settlement Type', 'Status', 'Created By', 'Approved By']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, ret in enumerate(returns, 2):
        ws.cell(row=row, column=1, value=ret.pr_number)
        ws.cell(row=row, column=2, value=ret.company.company_name)
        ws.cell(row=row, column=3, value=ret.return_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=4, value=ret.get_return_type_display())
        ws.cell(row=row, column=5, value=ret.get_return_reason_display())
        ws.cell(row=row, column=6, value=float(ret.total_amount))
        ws.cell(row=row, column=7, value=ret.get_settlement_type_display())
        ws.cell(row=row, column=8, value=ret.get_status_display())
        ws.cell(row=row, column=9, value=ret.created_by.get_full_name())
        ws.cell(row=row, column=10, value=ret.approved_by.get_full_name() if ret.approved_by else '-')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=purchase_returns_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


@login_required
def print_purchase_return_pdf(request, pk):
    """Generate PDF for Purchase Return matching exact physical form template"""
    
    # Access control
    if request.user.user_type not in ['admin', 'office']:
        messages.error(request, 'Access denied.')
        return redirect('dashboard:home')
    
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from business.models import DistributorProfile
    import io
    
    purchase_return = get_object_or_404(
        PurchaseReturn.objects.select_related('company', 'created_by'), 
        pk=pk
    )
    
    # Get business profile
    business = DistributorProfile.get_active()
    
    # Get items
    items = purchase_return.items.select_related('product').all()
    
    # Determine template based on return_type
    if purchase_return.return_type == 'expired':
        template_title = "EXPIRED RETURN NOTE"
        quantity_header = "EXPIRED BOTS<br/>QUANTITY"
        # Shop price formula (marked_price - shop_discount%)
        invoice_formula = "C=(B/100*90)"
        # Company price formula (shop_price - company_discount%)
        distributor_formula = "D=(C/100*77)"
    else:  # damaged
        template_title = "GAS OUT RETURN NOTE"
        quantity_header = "GAS OUT BOTS<br/>QUANTITY"
        # Shop price formula (marked_price - shop_discount%)
        invoice_formula = "C=(B/100*90)"
        # Company price formula (shop_price - company_discount%)
        distributor_formula = "D=(C/100*77)"
    
    # Create PDF
    buffer = io.BytesIO()
    
    # Set meaningful document title
    pdf_title = f"{template_title} - {purchase_return.pr_number}"
    
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=20*mm,
        leftMargin=20*mm,
        topMargin=15*mm,
        bottomMargin=15*mm,
        title=pdf_title
    )
    
    elements = []
    
    # ========== HEADER ==========
    header_style = ParagraphStyle('Header', fontSize=16, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=2)
    subheader_style = ParagraphStyle('SubHeader', fontSize=13, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=8)
    
    elements.append(Paragraph("MAX BEVERAGERS (PVT)LTD", header_style))
    elements.append(Paragraph(template_title, subheader_style))
    
    # ========== INFO SECTION ==========
    info_style = ParagraphStyle('Info', fontSize=9, leading=11)
    
    info_data = [
        [
            f"NAME OF DISTRIBUTOR :",
            f"{business.business_name}",
            f"DATE: {purchase_return.return_date.strftime('%d/%m/%Y')}"
        ],
        [
            f"AREA :",
            f"{business.city or 'Colombo'}",
            ""
        ],
        [
            f"MONTH :",
            f"{purchase_return.return_date.strftime('%B %Y')}",
            ""
        ]
    ]
    
    info_table = Table(info_data, colWidths=[45*mm, 75*mm, 50*mm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 3*mm))
    
    # ========== PRODUCT TABLE ==========
    # Get ALL products dynamically from database following display_order
    from collections import OrderedDict
    
    # Get all products from the company ordered by display_order (which groups by size)
    all_products = Product.objects.filter(
        company=purchase_return.company,
        is_active=True
    ).order_by('display_order', 'product_name')
    
    # Group products by size while maintaining display_order
    product_template = OrderedDict()
    for product in all_products:
        size = product.size
        if size not in product_template:
            product_template[size] = []
        product_template[size].append(product.product_name.upper())
    
    # Create lookup for returned items
    items_dict = {}
    for item in items:
        key = (item.product.size, item.product.product_name.upper())
        items_dict[key] = item
    
    # Build table without size column
    table_data = []
    
    # Header row - single PRODUCT RANGE column
    header_font = ParagraphStyle('Hdr', fontSize=7, alignment=TA_CENTER, fontName='Helvetica-Bold', leading=8)
    table_data.append([
        Paragraph("<b>PRODUCT RANGE</b>", header_font),
        Paragraph(f"<b>{quantity_header}</b><br/>(A)", header_font),
        Paragraph("<b>MARKED PRICE<br/>ON BOTTLE<br/>(B)</b>", header_font),
        Paragraph(f"<b>INVOICE PRICE<br/>{invoice_formula}</b>", header_font),
        Paragraph(f"<b>DISTRIBUTOR<br/>PRICE<br/>{distributor_formula}</b>", header_font),
        Paragraph("<b>TOTAL<br/>AMOUNT<br/>(E=A*D)</b>", header_font)
    ])
    
    total_amount = Decimal('0')
    cell_style = ParagraphStyle('Cell', fontSize=8, leading=9)
    cell_right = ParagraphStyle('CellR', fontSize=8, alignment=TA_RIGHT, leading=9)
    cell_center = ParagraphStyle('CellC', fontSize=8, alignment=TA_CENTER, leading=9)
    
    # Loop through sizes in the order they appear in product_template (preserves display_order)
    for size, products_in_size in product_template.items():
        # Add all products in this size (no size column)
        for product_name in products_in_size:
            key = (size, product_name)
            item = items_dict.get(key)
            
            if item:
                qty = str(item.quantity)
                marked = f"{item.marked_price:,.2f}"
                invoice = f"{item.invoice_price:,.2f}"
                distributor = f"{item.unit_price:,.2f}"
                amount = f"{item.line_total:,.2f}"
                total_amount += item.line_total
            else:
                qty = marked = invoice = distributor = amount = ''
            
            table_data.append([
                Paragraph(f"{size} {product_name}", cell_style),  # Size + product name in one column
                Paragraph(qty, cell_center),
                Paragraph(marked, cell_right),
                Paragraph(invoice, cell_right),
                Paragraph(distributor, cell_right),
                Paragraph(amount, cell_right)
            ])
    
    # Total row
    table_data.append([
        Paragraph("<b>TOTAL AMOUNT</b>", ParagraphStyle('Tot', fontSize=9, fontName='Helvetica-Bold', alignment=TA_CENTER)),
        '', '', '', '',
        Paragraph(f"<b>{total_amount:,.2f}</b>", ParagraphStyle('TotVal', fontSize=9, fontName='Helvetica-Bold', alignment=TA_RIGHT))
    ])
    
    # Create product table - 6 columns instead of 7
    product_table = Table(table_data, colWidths=[50*mm, 22*mm, 25*mm, 25*mm, 25*mm, 28*mm])
    
    # Build table style without rowspans
    table_style = [
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('SPAN', (0, -1), (4, -1)),  # Span total label across first 5 columns
    ]
    
    product_table.setStyle(TableStyle(table_style))
    elements.append(product_table)
    elements.append(Spacer(1, 3*mm))
    
    # ========== FOR OFFICE USE ==========
    office_data = [[Paragraph("<b>FOR OFFICE USE</b>", ParagraphStyle('Office', fontSize=9, fontName='Helvetica-Bold'))]]
    office_table = Table(office_data, colWidths=[165*mm], rowHeights=[6*mm])
    office_table.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(office_table)
    
    # Empty office space
    office_space = Table([['']], colWidths=[165*mm], rowHeights=[15*mm])
    office_space.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(office_space)
    elements.append(Spacer(1, 5*mm))
    
    # ========== SIGNATURE SECTION ==========
    # 3-column layout matching React template
    sig_style = ParagraphStyle('Sig', fontSize=8)
    
    sig_data = [
        [
            'DISTRIBUTOR SIGNATURE',
            'DRIVER NAME..........................',
            'LORRY NO.............................'
        ],
        [
            '....................................',
            'SIGNATURE............................',
            'TIME..................................'
        ]
    ]
    
    sig_table = Table(sig_data, colWidths=[55*mm, 55*mm, 55*mm], rowHeights=[8*mm, 8*mm])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('VALIGN', (0, 0), (-1, -1), 'BOTTOM'),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1, 3*mm))
    
    # ========== COMMENTS ==========
    comments_header = Table([[Paragraph("<b>COMMENTS</b>", ParagraphStyle('ComH', fontSize=9, fontName='Helvetica-Bold'))]], 
                            colWidths=[165*mm], rowHeights=[6*mm])
    comments_header.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(comments_header)
    
    comments_space = Table([['']], colWidths=[165*mm], rowHeights=[25*mm])
    comments_space.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
    ]))
    elements.append(comments_space)
    
    # Build PDF
    doc.build(elements)
    
    # Return response
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f'{"Expired" if purchase_return.return_type == "expired" else "GasOut"}_Return_{purchase_return.pr_number}.pdf'
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response
