"""
FOC (Free of Charge) Value Usage Account Views
Tracks FOC value received from companies vs FOC value given to shops
"""
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Sum, Q, Count, F, Value, DecimalField
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta

from products.models import FOCValueAccount, FOCValueTransaction, Company, Product
from shops.models import Shop


@login_required
def foc_dashboard(request):
    """FOC Value Usage Dashboard - Overview of all companies
    
    Admin/Office: See all FOC data across all companies and reps
    Sales Rep: See only their own FOC transactions
    """
    
    # Get all FOC accounts with related company
    accounts = FOCValueAccount.objects.select_related('company').all()
    
    # Filter transactions based on user type for transaction list only
    if request.user.is_sales_rep:
        # Sales reps see only their own transactions in the list
        transactions_filter = Q(sales_rep=request.user)
    else:
        # Admin/Office see all transactions
        transactions_filter = Q()
    
    # Calculate summary statistics - Same for all users (show overall company stats)
    total_foc_received = sum(acc.total_foc_received_value for acc in accounts)
    total_foc_given = sum(acc.total_foc_given_value for acc in accounts)
    net_foc_value = total_foc_received - total_foc_given
    
    # Get average utilization
    utilizations = [acc.foc_utilization_percentage for acc in accounts if acc.total_foc_received_value > 0]
    avg_utilization = sum(utilizations) / len(utilizations) if utilizations else 0
    
    # Get recent transactions (last 30 days)
    thirty_days_ago = timezone.now() - timedelta(days=30)
    recent_transactions = FOCValueTransaction.objects.filter(
        transactions_filter,
        transaction_date__gte=thirty_days_ago,
        is_archived=False
    ).select_related(
        'foc_account__company', 'product', 'shop', 'sales_rep',
        'purchase_item__purchase', 'bill_item__bill', 'return_item__return_ref'
    ).order_by('-transaction_date')[:50]
    
    # Transaction type breakdown
    txn_breakdown = FOCValueTransaction.objects.filter(
        transactions_filter,
        is_archived=False
    ).values('transaction_type').annotate(
        count=Count('id'),
        total_value=Sum('foc_value')
    )
    
    context = {
        'accounts': accounts,  # Show accounts to all users
        'total_foc_received': total_foc_received,
        'total_foc_given': total_foc_given,
        'net_foc_value': net_foc_value,
        'avg_utilization': avg_utilization,
        'recent_transactions': recent_transactions,
        'txn_breakdown': txn_breakdown,
        'is_sales_rep_view': request.user.is_sales_rep,
    }
    
    return render(request, 'sales/foc_dashboard.html', context)


@login_required
def foc_company_detail(request, company_id):
    """Detailed FOC transaction history for a specific company
    
    Admin/Office: See all transactions for the company
    Sales Rep: See only their own transactions for the company
    """
    
    company = get_object_or_404(Company, pk=company_id)
    
    # Get or create FOC account
    foc_account, created = FOCValueAccount.objects.get_or_create(
        company=company,
        defaults={'created_by': request.user}
    )
    
    # Get all transactions for this account
    transactions = foc_account.transactions.filter(
        is_archived=False
    ).select_related('product', 'shop', 'sales_rep').order_by('-transaction_date')
    
    # Filter by sales rep if user is a sales rep
    if request.user.is_sales_rep:
        transactions = transactions.filter(sales_rep=request.user)
    
    # Apply filters
    transaction_type = request.GET.get('type')
    if transaction_type:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    product_id = request.GET.get('product')
    if product_id:
        transactions = transactions.filter(product_id=product_id)
    
    date_from = request.GET.get('date_from')
    if date_from:
        transactions = transactions.filter(transaction_date__gte=date_from)
    
    date_to = request.GET.get('date_to')
    if date_to:
        transactions = transactions.filter(transaction_date__lte=date_to)
    
    # Product-wise summary
    product_summary = transactions.values(
        'product__product_name',
        'product__size',
        'product_id'
    ).annotate(
        total_foc_qty=Sum('foc_quantity'),
        total_foc_value=Sum('foc_value'),
        transaction_count=Count('id')
    ).order_by('-total_foc_value')
    
    # Monthly trend (last 12 months)
    twelve_months_ago = timezone.now() - timedelta(days=365)
    monthly_data = transactions.filter(
        transaction_date__gte=twelve_months_ago
    ).extra(
        select={'month': "DATE_TRUNC('month', transaction_date)"}
    ).values('month', 'transaction_type').annotate(
        total_value=Sum('foc_value')
    ).order_by('month')
    
    # Get all products for filter dropdown
    products = Product.objects.filter(
        company=company,
        is_active=True
    ).order_by('product_name')
    
    context = {
        'company': company,
        'account': foc_account,  # Template expects 'account' not 'foc_account'
        'foc_account': foc_account,  # Keep for backward compatibility
        'transactions': transactions[:100],  # Limit to recent 100
        'product_summary': product_summary,
        'monthly_data': monthly_data,
        'products': products,
        'selected_type': transaction_type,
        'selected_product': product_id,
        'date_from': date_from,
        'date_to': date_to,
    }
    
    return render(request, 'sales/foc_company_detail.html', context)


@login_required
def foc_product_report(request):
    """FOC value breakdown by product across all companies
    
    Admin/Office: See all FOC data across all reps
    Sales Rep: See only their own FOC data
    """
    
    # Get filter parameters
    company_id = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Base query
    query = FOCValueTransaction.objects.filter(is_archived=False)
    
    # Filter by sales rep if user is a sales rep
    if request.user.is_sales_rep:
        query = query.filter(sales_rep=request.user)
    
    # Apply filters with timezone-aware dates
    if company_id:
        query = query.filter(product__company_id=company_id)
    
    if start_date:
        from datetime import datetime
        start_date_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        query = query.filter(transaction_date__gte=start_date_dt)
    
    if end_date:
        from datetime import datetime
        end_date_dt = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        query = query.filter(transaction_date__lte=end_date_dt)
    
    # Get product-wise FOC summary with all required fields including returns
    product_summary = query.values(
        'product__product_name',
        'product__size',
        'product__marked_price',
        'product__company__company_name',
        'product__display_order',
        'product_id'
    ).annotate(
        foc_received_qty=Sum('foc_quantity', filter=Q(transaction_type='foc_received')),
        foc_received_value=Sum('foc_value', filter=Q(transaction_type='foc_received')),
        foc_given_qty=Sum('foc_quantity', filter=Q(transaction_type='foc_given')),
        foc_given_value=Sum('foc_value', filter=Q(transaction_type='foc_given')),
        implicit_foc_value=Sum('foc_value', filter=Q(transaction_type='implicit_foc')),
        foc_returned_qty=Sum('foc_quantity', filter=Q(transaction_type='return_foc_restored')),
        foc_returned_value=Sum('foc_value', filter=Q(transaction_type='return_foc_restored')),
        net_foc_value=(
            Coalesce(Sum('foc_value', filter=Q(transaction_type='foc_received')), Value(0, output_field=DecimalField())) - 
            Coalesce(Sum('foc_value', filter=Q(transaction_type='foc_given')), Value(0, output_field=DecimalField())) -
            Coalesce(Sum('foc_value', filter=Q(transaction_type='implicit_foc')), Value(0, output_field=DecimalField())) +
            Coalesce(Sum('foc_value', filter=Q(transaction_type='return_foc_restored')), Value(0, output_field=DecimalField()))
        )
    ).order_by('product__display_order', 'product__product_name')
    
    # Build product_data list with proper field names for template
    product_data = []
    totals = {
        'total_foc_received_qty': Decimal('0'),
        'total_foc_received_value': Decimal('0'),
        'total_foc_given_qty': Decimal('0'),
        'total_foc_given_value': Decimal('0'),
        'total_implicit_foc_value': Decimal('0'),
        'total_foc_returned_qty': Decimal('0'),
        'total_foc_returned_value': Decimal('0'),
        'net_total': Decimal('0'),
    }
    
    for item in product_summary:
        # Build product display name
        product_name = f"{item['product__product_name']} - {item['product__size']}"
        
        # Replace None with 0 and build clean dict
        foc_received_qty = item['foc_received_qty'] or Decimal('0')
        foc_received_value = item['foc_received_value'] or Decimal('0')
        foc_given_qty = item['foc_given_qty'] or Decimal('0')
        foc_given_value = item['foc_given_value'] or Decimal('0')
        implicit_foc_value = item['implicit_foc_value'] or Decimal('0')
        foc_returned_qty = item['foc_returned_qty'] or Decimal('0')
        foc_returned_value = item['foc_returned_value'] or Decimal('0')
        net_foc_value = item['net_foc_value'] or Decimal('0')
        
        product_data.append({
            'product': product_name,
            'company': item['product__company__company_name'],
            'foc_received_qty': foc_received_qty,
            'foc_received_value': foc_received_value,
            'foc_given_qty': foc_given_qty,
            'foc_given_value': foc_given_value,
            'implicit_foc_value': implicit_foc_value,
            'foc_returned_qty': foc_returned_qty,
            'foc_returned_value': foc_returned_value,
            'net_foc_value': net_foc_value,
        })
        
        # Update totals
        totals['total_foc_received_qty'] += foc_received_qty
        totals['total_foc_received_value'] += foc_received_value
        totals['total_foc_given_qty'] += foc_given_qty
        totals['total_foc_given_value'] += foc_given_value
        totals['total_implicit_foc_value'] += implicit_foc_value
        totals['total_foc_returned_qty'] += foc_returned_qty
        totals['total_foc_returned_value'] += foc_returned_value
        totals['net_total'] += net_foc_value
    
    # Get all companies for filter dropdown
    from products.models import Company
    companies = Company.objects.all().order_by('company_name')
    
    context = {
        'product_data': product_data,
        'totals': totals,
        'companies': companies,
    }
    
    return render(request, 'sales/foc_product_report.html', context)


@login_required
def foc_sales_rep_report(request):
    """FOC given breakdown by sales representative
    
    Admin/Office: See data for all sales reps
    Sales Rep: See only their own data
    """
    
    # Get filter parameters
    company_id = request.GET.get('company')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    # Default to last 30 days if not specified
    if not start_date:
        start_date_dt = timezone.now() - timedelta(days=30)
    else:
        from datetime import datetime
        start_date_dt = timezone.make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
    
    if not end_date:
        end_date_dt = timezone.now()
    else:
        from datetime import datetime
        end_date_dt = timezone.make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    
    # Base query with timezone-aware date filtering - include returns
    query = FOCValueTransaction.objects.filter(
        transaction_type__in=['foc_given', 'implicit_foc', 'return_foc_restored'],
        transaction_date__gte=start_date_dt,
        transaction_date__lte=end_date_dt,
        is_archived=False,
        sales_rep__isnull=False
    )
    
    # Apply company filter if specified
    if company_id:
        query = query.filter(product__company_id=company_id)
    
    # Sales rep summary - returns reduce FOC used (properly handle null values)
    rep_summary = query.values(
        'sales_rep__username',
        'sales_rep__first_name',
        'sales_rep__last_name',
        'sales_rep_id'
    ).annotate(
        foc_given_qty=Sum('foc_quantity', filter=Q(transaction_type='foc_given')),
        foc_given_value=Sum('foc_value', filter=Q(transaction_type='foc_given')),
        implicit_foc_value=Sum('foc_value', filter=Q(transaction_type='implicit_foc')),
        foc_returned_qty=Sum('foc_quantity', filter=Q(transaction_type='return_foc_restored')),
        foc_returned_value=Sum('foc_value', filter=Q(transaction_type='return_foc_restored')),
        total_foc_used=(
            Coalesce(Sum('foc_value', filter=Q(transaction_type__in=['foc_given', 'implicit_foc'])), Value(0, output_field=DecimalField())) - 
            Coalesce(Sum('foc_value', filter=Q(transaction_type='return_foc_restored')), Value(0, output_field=DecimalField()))
        ),
        bills_count=Count('bill_item_id', distinct=True)
    ).order_by('-total_foc_used')
    
    # Build sales_rep_data list with proper field names for template
    sales_rep_data = []
    totals = {
        'total_foc_given_qty': Decimal('0'),
        'total_foc_given_value': Decimal('0'),
        'total_implicit_foc_value': Decimal('0'),
        'total_foc_returned_qty': Decimal('0'),
        'total_foc_returned_value': Decimal('0'),
        'total_foc_used': Decimal('0'),
        'total_bills_count': 0,
    }
    
    for item in rep_summary:
        # Build sales rep name
        rep_name = f"{item['sales_rep__first_name']} {item['sales_rep__last_name']}"
        if not rep_name.strip():
            rep_name = item['sales_rep__username']
        
        # Replace None with 0
        foc_given_qty = item['foc_given_qty'] or Decimal('0')
        foc_given_value = item['foc_given_value'] or Decimal('0')
        implicit_foc_value = item['implicit_foc_value'] or Decimal('0')
        foc_returned_qty = item['foc_returned_qty'] or Decimal('0')
        foc_returned_value = item['foc_returned_value'] or Decimal('0')
        total_foc_used = item['total_foc_used'] or Decimal('0')
        bills_count = item['bills_count'] or 0
        
        # Calculate average FOC per bill
        avg_foc_per_bill = (total_foc_used / bills_count) if bills_count > 0 else Decimal('0')
        
        sales_rep_data.append({
            'sales_rep': rep_name,
            'foc_given_qty': foc_given_qty,
            'foc_given_value': foc_given_value,
            'implicit_foc_value': implicit_foc_value,
            'foc_returned_qty': foc_returned_qty,
            'foc_returned_value': foc_returned_value,
            'total_foc_used': total_foc_used,
            'bills_count': bills_count,
            'avg_foc_per_bill': avg_foc_per_bill,
        })
        
        # Update totals
        totals['total_foc_given_qty'] += foc_given_qty
        totals['total_foc_given_value'] += foc_given_value
        totals['total_implicit_foc_value'] += implicit_foc_value
        totals['total_foc_returned_qty'] += foc_returned_qty
        totals['total_foc_returned_value'] += foc_returned_value
        totals['total_foc_used'] += total_foc_used
        totals['total_bills_count'] += bills_count
    
    # Calculate overall average
    totals['avg_foc_per_bill'] = (totals['total_foc_used'] / totals['total_bills_count']) if totals['total_bills_count'] > 0 else Decimal('0')
    
    # Get all companies for filter dropdown
    from products.models import Company
    companies = Company.objects.all().order_by('company_name')
    
    context = {
        'sales_rep_data': sales_rep_data,
        'totals': totals,
        'companies': companies,
        'date_from': start_date if isinstance(start_date, str) else start_date_dt.strftime('%Y-%m-%d'),
        'date_to': end_date if isinstance(end_date, str) else end_date_dt.strftime('%Y-%m-%d'),
    }
    
    return render(request, 'sales/foc_sales_rep_report.html', context)


@login_required
def reset_foc_account(request, company_id):
    """Reset FOC account with historical archival"""
    if request.user.user_type != 'admin':
        messages.error(request, 'Only administrators can reset FOC accounts.')
        return redirect('sales:foc_dashboard')
    
    company = get_object_or_404(Company, pk=company_id)
    foc_account = get_object_or_404(FOCValueAccount, company=company)
    
    if request.method == 'POST':
        try:
            # Get reset notes
            reset_notes = request.POST.get('reset_notes', '')
            
            # Archive all current transactions
            archived_count = foc_account.transactions.filter(
                is_archived=False
            ).update(is_archived=True)
            
            # Create adjustment transaction to close current period
            if foc_account.net_foc_value != 0:
                FOCValueTransaction.objects.create(
                    foc_account=foc_account,
                    transaction_type='adjustment',
                    transaction_date=timezone.now(),
                    product=Product.objects.filter(company=company).first(),  # Use any product from company
                    foc_quantity=0,
                    shop_price_at_time=0,
                    foc_value=-foc_account.net_foc_value,  # Reverse to zero out
                    reference_number=f'RESET-{timezone.now().strftime("%Y%m%d")}',
                    notes=f'Account reset - Closing balance: Rs.{foc_account.net_foc_value:,.2f}. {reset_notes}',
                    created_by=request.user,
                    is_archived=True  # Immediately archive the closing entry
                )
            
            # Update opening balances to current totals
            foc_account.opening_foc_received_value = foc_account.total_foc_received_value
            foc_account.opening_foc_given_value = foc_account.total_foc_given_value
            foc_account.opening_date = timezone.localdate()
            foc_account.opening_notes = f'Reset by {request.user.username}. Previous balance archived. {reset_notes}'
            foc_account.save()
            
            # Recalculate balances (should now only include new opening balances)
            foc_account.update_balance()
            
            messages.success(
                request,
                f'FOC account for {company.company_name} reset successfully. '
                f'{archived_count} transactions archived.'
            )
            
        except Exception as e:
            messages.error(request, f'Error resetting account: {str(e)}')
        
        return redirect('sales:foc_company_detail', company_id=company_id)
    
    # GET request - show confirmation page
    transactions = foc_account.transactions.filter(is_archived=False)
    
    context = {
        'company': company,
        'foc_account': foc_account,
        'transaction_count': transactions.count(),
        'transactions': transactions.order_by('-transaction_date')[:10],  # Show last 10
    }
    
    return render(request, 'sales/foc_reset_confirm.html', context)


@login_required
def export_foc_transactions(request, company_id):
    """Export FOC transactions to Excel
    
    Admin/Office: Export all transactions
    Sales Rep: Export only their own transactions
    """
    
    company = get_object_or_404(Company, pk=company_id)
    foc_account = get_object_or_404(FOCValueAccount, company=company)
    
    # Get all non-archived transactions
    transactions = foc_account.transactions.filter(
        is_archived=False
    ).select_related('product', 'shop', 'sales_rep').order_by('transaction_date')
    
    # Filter by sales rep if user is a sales rep
    if request.user.is_sales_rep:
        transactions = transactions.filter(sales_rep=request.user)
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FOC Transactions"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Headers
    headers = [
        'Transaction #', 'Date', 'Type', 'Product', 'FOC Qty', 
        'Shop Price', 'FOC Value', 'Reference', 'Shop', 'Sales Rep', 'Notes'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Data rows
    for row, txn in enumerate(transactions, 2):
        ws.cell(row, 1, txn.transaction_number)
        from django.utils import timezone as tz
        ws.cell(row, 2, tz.localtime(txn.transaction_date).strftime('%Y-%m-%d %H:%M'))
        ws.cell(row, 3, txn.get_transaction_type_display())
        ws.cell(row, 4, txn.product.product_name)
        ws.cell(row, 5, float(txn.foc_quantity))
        ws.cell(row, 6, float(txn.shop_price_at_time))
        ws.cell(row, 7, float(txn.foc_value))
        ws.cell(row, 8, txn.reference_number)
        ws.cell(row, 9, txn.shop.shop_name if txn.shop else '')
        ws.cell(row, 10, txn.sales_rep.username if txn.sales_rep else '')
        ws.cell(row, 11, txn.notes or '')
    
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 25
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 40
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="FOC_Transactions_{company.company_name}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb.save(response)
    return response
